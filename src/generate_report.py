"""
Automated Sales Report Generator

Reads a CSV file with sales data, cleans it, calculates business KPIs,
and generates an Excel report with summary tables and charts.

Usage:
    python src/generate_report.py --input data/sales_sample.csv --output output/sales_report.xlsx
"""

from __future__ import annotations

import argparse
import sys
import warnings
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = {
    "date",
    "client",
    "region",
    "product",
    "units",
    "unit_price",
    "status",
}

ALLOWED_STATUS = frozenset({"paid", "pending", "cancelled"})

# Rows (incl. header) scanned for column width on large "Clean Data" dumps — avoids O(rows×cols) openpyxl walks.
_CLEAN_DATA_WIDTH_SAMPLE_ROWS = 1000


def _normalize_column_names(df: pd.DataFrame) -> None:
    """Strip headers and lower-case them so exports like 'Date' still match."""
    cols = [str(c).strip().lower() for c in df.columns]
    dup_counts = Counter(cols)
    dupes = sorted(c for c, n in dup_counts.items() if n > 1)
    if dupes:
        raise ValueError(f"Duplicate column names after normalizing case/spaces: {dupes}")
    df.columns = cols


def _validate_status(series: pd.Series) -> None:
    """Ensure every row uses an allowed status (README contract)."""
    invalid_mask = ~series.isin(ALLOWED_STATUS)
    if not invalid_mask.any():
        return
    bad = series[invalid_mask]
    labels = []
    for v in bad.dropna().unique():
        labels.append(repr(str(v)))
    if bad.isna().any():
        labels.append("<missing/blank>")
    seen = sorted(set(labels))
    raise ValueError(f"Invalid status value(s) in CSV (allowed: {sorted(ALLOWED_STATUS)}): {', '.join(seen)}")


def _strip_string_column(series: pd.Series) -> pd.Series:
    """Coerce to nullable string so .str works with numeric columns and NaN."""
    return series.astype("string").str.strip()


def load_sales_data(input_path: Path, *, encoding: str = "utf-8") -> pd.DataFrame:
    """Load, validate and clean the raw sales CSV."""
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    df = pd.read_csv(input_path, encoding=encoding)
    _normalize_column_names(df)
    missing_columns = REQUIRED_COLUMNS - set(df.columns)
    if missing_columns:
        raise ValueError(f"Missing required columns: {sorted(missing_columns)}")

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["client"] = _strip_string_column(df["client"])
    df["region"] = _strip_string_column(df["region"])
    df["product"] = _strip_string_column(df["product"])
    df["units"] = pd.to_numeric(df["units"], errors="coerce").fillna(0).astype(int)
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0.0)
    df["status"] = _strip_string_column(df["status"]).str.lower()
    _validate_status(df["status"])
    df["revenue"] = df["units"] * df["unit_price"]
    df["month"] = df["date"].dt.to_period("M").astype(str)

    n_before = len(df)
    df = df.dropna(subset=["date"])
    dropped_dates = n_before - len(df)
    if dropped_dates:
        warnings.warn(
            f"Dropped {dropped_dates} row(s) with invalid or missing dates.",
            UserWarning,
            stacklevel=3,
        )
    return df


def build_summary_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Create the summary tables used in the Excel report."""
    valid_sales = df[df["status"].isin(["paid", "pending"])]
    paid_sales = df[df["status"] == "paid"]

    return {
        "KPIs": pd.DataFrame(
            {
                "metric": [
                    "Total revenue",
                    "Paid revenue",
                    "Pending revenue",
                    "Cancelled revenue",
                    "Total units sold",
                    "Number of clients",
                ],
                "value": [
                    valid_sales["revenue"].sum(),
                    paid_sales["revenue"].sum(),
                    df[df["status"] == "pending"]["revenue"].sum(),
                    df[df["status"] == "cancelled"]["revenue"].sum(),
                    valid_sales["units"].sum(),
                    df["client"].nunique(),
                ],
            }
        ),
        "Monthly Revenue": valid_sales.groupby("month", as_index=False)["revenue"].sum(),
        "Revenue by Region": valid_sales.groupby("region", as_index=False)["revenue"]
        .sum()
        .sort_values("revenue", ascending=False),
        "Revenue by Product": valid_sales.groupby("product", as_index=False)["revenue"]
        .sum()
        .sort_values("revenue", ascending=False),
        "Top Clients": paid_sales.groupby("client", as_index=False)["revenue"]
        .sum()
        .sort_values("revenue", ascending=False)
        .head(5),
        "Clean Data": df,
    }


def _style_header_row(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _apply_column_widths_bounded(ws, *, last_row: int) -> None:
    """Set column widths from values in rows 1..last_row (avoids ws.columns full scans)."""
    last_row = max(1, min(last_row, ws.max_row or 1))
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        max_length = 0
        for row in range(1, last_row + 1):
            val = ws.cell(row=row, column=col).value
            max_length = max(max_length, len(str(val if val is not None else "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 35)


def style_worksheet(ws, *, lightweight: bool = False) -> None:
    """Apply readable formatting. Use lightweight=True for large raw-data sheets (no per-cell body borders)."""
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))
    _style_header_row(ws)

    if lightweight:
        sample_end = min(ws.max_row or 1, _CLEAN_DATA_WIDTH_SAMPLE_ROWS)
        _apply_column_widths_bounded(ws, last_row=sample_end)
        ws.freeze_panes = "A2"
        return

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    _apply_column_widths_bounded(ws, last_row=ws.max_row or 1)
    ws.freeze_panes = "A2"


def add_charts(workbook_path: Path) -> None:
    """Add Excel charts to the generated workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)

    ws = wb["Monthly Revenue"]
    chart = LineChart()
    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")

    ws = wb["Revenue by Product"]
    chart = BarChart()
    chart.title = "Revenue by Product"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Product"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")

    try:
        wb.save(workbook_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot save charts to '{workbook_path.resolve()}'. Close the workbook in Excel "
            "or choose a different --output path."
        ) from exc


def _output_write_permission_message(path: Path) -> str:
    return (
        f"Cannot write '{path.resolve()}'. If the file is open in Excel (or another program), "
        "close it or use a different --output path."
    )


def generate_report(input_path: Path, output_path: Path, *, encoding: str = "utf-8") -> None:
    """Generate the final Excel report."""
    df = load_sales_data(input_path, encoding=encoding)
    tables = build_summary_tables(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, table in tables.items():
                table.to_excel(writer, index=False, sheet_name=sheet_name)
                style_worksheet(
                    writer.book[sheet_name],
                    lightweight=(sheet_name == "Clean Data"),
                )
        add_charts(output_path)
    except PermissionError as exc:
        raise PermissionError(_output_write_permission_message(output_path)) from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an automated sales Excel report.")
    parser.add_argument("--input", type=Path, default=Path("data/sales_sample.csv"))
    parser.add_argument("--output", type=Path, default=Path("output/sales_report.xlsx"))
    parser.add_argument(
        "--encoding",
        default="utf-8",
        help="Text encoding for the input CSV (e.g. utf-8, utf-8-sig, cp1252).",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        generate_report(args.input, args.output, encoding=args.encoding)
    except PermissionError as exc:
        print(exc, file=sys.stderr)
        raise SystemExit(1) from exc
    print(f"Report generated successfully: {args.output}")
