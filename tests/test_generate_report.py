"""Tests for sales report generation."""

from __future__ import annotations

import textwrap
from pathlib import Path
import pandas as pd
import pytest
from openpyxl import Workbook
from src.generate_report import (
    _output_write_permission_message,
    generate_report,
    load_sales_data,
    style_worksheet,
)


def _write_csv(path: Path, content: str) -> Path:
    path.write_text(textwrap.dedent(content).lstrip(), encoding="utf-8")
    return path


def test_load_sales_data_normalizes_headers_and_numeric_client(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path / "in.csv",
        """
        Date,Client,Region,Product,Units,Unit_Price,Status
        2026-01-01,12345,Madrid,Widget,2,10.5,paid
        """,
    )
    df = load_sales_data(csv_path)
    assert len(df) == 1
    assert df["client"].iloc[0] == "12345"
    assert df["status"].iloc[0] == "paid"


def test_load_sales_data_invalid_status_raises(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path / "bad.csv",
        """
        date,client,region,product,units,unit_price,status
        2026-01-01,A,M,P,1,1,payed
        """,
    )
    with pytest.raises(ValueError, match="Invalid status"):
        load_sales_data(csv_path)


def test_load_sales_data_warns_on_dropped_dates(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path / "dates.csv",
        """
        date,client,region,product,units,unit_price,status
        2026-01-01,A,M,P,1,1,paid
        not-a-date,B,M,P,1,1,paid
        """,
    )
    with pytest.warns(UserWarning, match="Dropped 1 row"):
        df = load_sales_data(csv_path)
    assert len(df) == 1


def test_generate_report_creates_workbook(tmp_path: Path) -> None:
    csv_path = _write_csv(
        tmp_path / "in.csv",
        """
        date,client,region,product,units,unit_price,status
        2026-01-01,Acme,Madrid,Widget,1,10,paid
        """,
    )
    out = tmp_path / "report.xlsx"
    generate_report(csv_path, out)
    assert out.is_file()
    xl = pd.ExcelFile(out)
    assert "Clean Data" in xl.sheet_names
    assert "KPIs" in xl.sheet_names


def test_style_worksheet_lightweight_skips_body_borders() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Data"
    for r in range(1, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")

    style_worksheet(ws, lightweight=True)
    bottom = ws["A2"].border.bottom
    assert getattr(bottom, "style", None) != "thin"


def test_style_worksheet_full_applies_body_border() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="h")
    ws.cell(row=2, column=1, value="x")

    style_worksheet(ws, lightweight=False)
    assert ws["A2"].border.bottom is not None
    assert ws["A2"].border.bottom.style == "thin"


def test_generate_report_permission_error_wraps_message(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    csv_path = _write_csv(
        tmp_path / "in.csv",
        """
        date,client,region,product,units,unit_price,status
        2026-01-01,A,M,P,1,1,paid
        """,
    )
    out = tmp_path / "out.xlsx"

    def boom(*_a, **_k):
        raise PermissionError(13, "Access denied")

    monkeypatch.setattr(pd, "ExcelWriter", boom)
    with pytest.raises(PermissionError) as ei:
        generate_report(csv_path, out)
    assert "Cannot write" in str(ei.value)
    assert str(out.resolve()) in str(ei.value) or "out.xlsx" in str(ei.value)


def test_generate_report_permission_on_add_charts(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    import src.generate_report as mod

    csv_path = _write_csv(
        tmp_path / "in.csv",
        """
        date,client,region,product,units,unit_price,status
        2026-01-01,A,M,P,1,1,paid
        """,
    )
    out = tmp_path / "out.xlsx"

    def fail_charts(_path: Path) -> None:
        raise PermissionError(13, "locked")

    monkeypatch.setattr(mod, "add_charts", fail_charts)
    with pytest.raises(PermissionError, match="Cannot write"):
        generate_report(csv_path, out)


def test_output_write_permission_message_contains_path() -> None:
    p = Path("output") / "sales_report.xlsx"
    msg = _output_write_permission_message(p)
    assert "sales_report.xlsx" in msg
    assert "Excel" in msg
